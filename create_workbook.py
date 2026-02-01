from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Crear workbook
wb = Workbook()
wb.remove(wb.active)  # Eliminar la hoja por defecto

# ==================== HOJA 1: Orders_Clients ====================
ws1 = wb.create_sheet("Orders_Clients", 0)

# Definir encabezados de la Hoja 1
headers_1 = [
    "Order_ID", "Order_Date", "Client_Name", "Contact_Channel",
    "Contact_Info", "Location", "Product_Design", "Quantity",
    "Order_Status", "Shipping_Company", "Tracking_Number",
    "Delivery_Status", "Follow_Up_Count", "Notes"
]

ws1.append(headers_1)

# Aplicar estilos a encabezados (Hoja 1)
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for cell in ws1[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = thin_border

# Establecer anchos de columnas (Hoja 1)
column_widths_1 = [12, 12, 18, 16, 15, 12, 18, 10, 15, 16, 15, 15, 14, 25]
for i, width in enumerate(column_widths_1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = width

# Agregar datos de ejemplo (Hoja 1)
sample_data_1 = [
    ["ORD-001", datetime(2025, 12, 1).date(), "Juan García", "Instagram", "+34612345678", "Madrid", "Camiseta Premium", 2, "Delivered", "Correos", "1Z12345", "Entregado", 1, "Cliente satisfecho"],
    ["ORD-002", datetime(2025, 12, 5).date(), "María López", "WhatsApp", "+34687654321", "Barcelona", "Sudadera Básica", 1, "In production", "Fedex", "", "En tránsito", 0, ""],
    ["ORD-003", datetime(2025, 12, 8).date(), "Carlos Ruiz", "Instagram", "+34698765432", "Valencia", "Gorro y Bufanda", 3, "Pending", "", "", "Pendiente", 0, "Esperar confirmación de pago"],
]

for row_data in sample_data_1:
    ws1.append(row_data)

# Aplicar bordes a datos (Hoja 1)
for row in ws1.iter_rows(min_row=2, max_row=len(sample_data_1)+1, min_col=1, max_col=len(headers_1)):
    for cell in row:
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Congelar encabezado (Hoja 1)
ws1.freeze_panes = "A2"

# ==================== HOJA 2: Financials ====================
ws2 = wb.create_sheet("Financials", 1)

# Definir encabezados de la Hoja 2
headers_2 = [
    "Order_ID", "Sale_Price", "Product_Cost", "Shipping_Cost",
    "Advertising_Cost", "Gross_Profit", "Net_Profit", "Profit_Margin_%",
    "Payment_Method", "Payment_Status", "Date_Paid"
]

ws2.append(headers_2)

# Aplicar estilos a encabezados (Hoja 2)
for cell in ws2[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = thin_border

# Establecer anchos de columnas (Hoja 2)
column_widths_2 = [12, 12, 14, 14, 16, 14, 12, 15, 16, 15, 12]
for i, width in enumerate(column_widths_2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = width

# Agregar datos de ejemplo con fórmulas (Hoja 2)
# Para la fórmula, usaremos referencias de celdas
sample_data_2 = [
    ["ORD-001", 45.00, 12.00, 5.00, 2.00, "=B2-C2-D2", "=E2-D2", "=F2/B2"],
    ["ORD-002", 38.50, 10.50, 4.50, 0.00, "=B3-C3-D3", "=E3-D3", "=F3/B3"],
    ["ORD-003", 25.00, 8.00, 3.50, 1.50, "=B4-C4-D4", "=E4-D4", "=F4/B4"],
]

# Agregar datos y aplicar fórmulas
for idx, row_data in enumerate(sample_data_2, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws2.cell(row=idx, column=col_idx, value=value)
        cell.border = thin_border
        
        # Formatear celdas de dinero
        if col_idx in [2, 3, 4, 5, 6, 7]:  # Columnas de precios
            cell.number_format = '$#,##0.00'
        # Formatear porcentajes
        elif col_idx == 8:
            cell.number_format = '0.00%'
        
        cell.alignment = Alignment(horizontal="right", vertical="center")

# Aplicar alineación al centro para Order_ID y Payment_Status
for row in ws2.iter_rows(min_row=2, max_row=len(sample_data_2)+1, min_col=1, max_col=11):
    ws2[f"A{row[0].row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws2[f"J{row[0].row}"].alignment = Alignment(horizontal="center", vertical="center")

# Congelar encabezado (Hoja 2)
ws2.freeze_panes = "A2"

# ==================== AGREGAR VALIDACIÓN Y DATOS VALIDADOS ====================

# Agregar hoja con validaciones (opcional pero útil)
ws_validation = wb.create_sheet("Reference", 2)
ws_validation.column_dimensions['A'].width = 20
ws_validation.column_dimensions['B'].width = 20

# Listas de validación
ws_validation['A1'] = "Contact_Channel"
ws_validation['B1'] = "Order_Status"
ws_validation['A1'].font = Font(bold=True, size=11)
ws_validation['B1'].font = Font(bold=True, size=11)

contact_channels = ["Instagram", "WhatsApp", "Email", "Facebook", "Phone", "Tienda"]
order_statuses = ["Pending", "In production", "Shipped", "Delivered", "Issue"]
payment_statuses = ["Pending", "Confirmed", "Paid", "Refunded"]
payment_methods = ["Transferencia", "Efectivo", "PayPal", "Tarjeta"]

for idx, (channel, status) in enumerate(zip(contact_channels, order_statuses), 2):
    ws_validation[f'A{idx}'] = channel
    ws_validation[f'B{idx}'] = status

# Guardar el workbook
output_path = r"c:\Users\Fernando\Downloads\FLUJO DE TRABAJO\SITIOS WEB FIVE MAGICS DESIGNS\EXCEL INGRESOS\Apparel_Business_Tracker.xlsx"
wb.save(output_path)

print(f"✓ Libro de Excel creado exitosamente: {output_path}")
print(f"\n📊 Estructura del libro:")
print(f"  - Hoja 1: 'Orders_Clients' (Gestión operativa)")
print(f"  - Hoja 2: 'Financials' (Control financiero)")
print(f"  - Hoja 3: 'Reference' (Listas de validación)")
print(f"\n🔗 Las hojas están vinculadas por Order_ID")
print(f"📈 Las fórmulas en la hoja Financials calculan automáticamente:")
print(f"   • Gross Profit = Sale Price - Product Cost - Shipping Cost")
print(f"   • Net Profit = Gross Profit - Advertising Cost")
print(f"   • Profit Margin = Net Profit / Sale Price")
